import os
import sys
import openpyxl
from duty_schedule.daily_duty_schedule import Daily_Duty_Schedule 

class Duty_Schedule_Reader():
	"""docstring for Duty_Schedule_Reader"""
	def __init__(self, duty_schedule_path):
		super(Duty_Schedule_Reader, self).__init__()
		self.duty_schedule_path = duty_schedule_path
		#
		try:
			self._excel_file = openpyxl.load_workbook(duty_schedule_path)
			self._sheet_names = self._excel_file.sheetnames
			if u'技术' in self._sheet_names:
				self._sheet_tech = self._excel_file[u'技术']
				self._sheet_tech_max_row = self._sheet_tech.max_row
				self._sheet_tech_max_column = self._sheet_tech.max_column
			else:
				print(u'没有找到名为‘技术’的工作表')
				os.system("pause");
				sys.exit()

		except Exception as e:
			raise e
		
		
	def __iter__(self):
		self._i = 1
		return self

	def __next__(self):
		if self._i < self._sheet_tech_max_row:
			self._i += 1 
			return Daily_Duty_Schedule(self._sheet_tech[self._i][0].value, self._sheet_tech[self._i][1].value, [self._sheet_tech[self._i][2].value, self._sheet_tech[self._i][3].value], [self._sheet_tech[self._i][4].value, self._sheet_tech[self._i][5].value], [self._sheet_tech[self._i][6].value, self._sheet_tech[self._i][7].value])
		elif self._i == self._sheet_tech_max_row:
			self._i += 1#表示次月的1号，星期7日为一周期，通宵班5个班为一周期
			return Daily_Duty_Schedule(self._sheet_tech[self._i-1][0].value+1, self._sheet_tech[self._i-7][1].value, [self._sheet_tech[self._i-5][2].value, None], [None, None], [None, None])
		else:
			raise StopIteration	

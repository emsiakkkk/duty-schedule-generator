import os
import sys
import openpyxl
from duty_schedule.monthly_duty_schedule import Monthly_Duty_Schedule 

class Duty_Schedule_Writer(object):
	"""docstring for Duty_Schedule_Writer"""
	def __init__(self, output_file, schedule : Monthly_Duty_Schedule):
		super(Duty_Schedule_Writer, self).__init__()
		if not output_file:
			print(u'值班表导入模板不能为空')
			os.system("pause");
			sys.exit()
		self._schedule = schedule
		self._output_file = output_file
		try:
			self._excel_file = openpyxl.load_workbook(self._output_file)
			self._sheet_names = self._excel_file.sheetnames
			if 'Sheet1' in self._sheet_names:
				self._sheet = self._excel_file['Sheet1']
			else:
				print(u'没有找到名为‘Sheet1’的工作表')
				os.system("pause");
				sys.exit()

		except Exception as e:
			raise e
	
	def write_duty_schedule_detail(self):
		for i in range(0, len(self._schedule.all_staffs)):
			#写入值班员名字
			self._sheet.cell(4, i + 3, self._schedule.all_staffs[i])
			# print(self._schedule.all_staffs[i])
			#print(self._schedule.staff(self._schedule.all_staffs[i]))
			# 获取当月指定值班员值班情况
			staff_schedule = self._schedule.staff(self._schedule.all_staffs[i])
			# print(staff_schedule)
		
			row = 0
			for date, s in staff_schedule.items():
				# 会导致通宵班押后一晚
				# if len(s) > 0:
				# 	self._sheet.cell(5 + row + s[0], i + 3, u'平台运维岗')
				# 	if len(s) == 2: #上午班、下午班直落
				# 		self._sheet.cell(5 + row + s[1], i + 3, u'平台运维岗')
				# 	if len(s) > 2: #上满三个班，不可能吧
				# 		pass
				
				if 0 in s:#上午班
					self._sheet.cell(5 + row + 0, i + 3, u'平台运维岗')

				if 1 in s:#下午班
					self._sheet.cell(5 + row + 1, i + 3, u'平台运维岗')

				# 填充日期应提前一日（忽略每月1日），否则会导致模板表中值班员姓名被替换为
				# 平台运维岗，上传时会触发duty_schedule_detail.people外键约束失败告警
				# if 2 in s and date[-2:] != "01":#通宵班
				if 2 in s and row >= 3:#通宵班,只跳过当月1日，但不跳过次月1日通宵班
					self._sheet.cell(5 + row + (-1), i + 3, u'平台运维岗')

				row += 3	

		self._excel_file.save(self._output_file)	

